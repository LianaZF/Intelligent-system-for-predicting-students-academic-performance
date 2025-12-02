from pathlib import Path
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.ensemble import (
    VotingClassifier, VotingRegressor, 
    RandomForestClassifier, RandomForestRegressor
)
from sklearn.neighbors import KNeighborsClassifier, KNeighborsRegressor
from sklearn.naive_bayes import GaussianNB
from sklearn.linear_model import LinearRegression
from sklearn.metrics import (
    accuracy_score, recall_score, r2_score,
    mean_absolute_error, mean_squared_error
)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
df = pd.read_csv(ROOT / "data" / "students_15000_CTGAN.csv")
df['risk_level'] = df['risk_category'].map({'В зоне риска': 0, 'Требует внимания': 1, 'Успешный студент': 2})

FEATURES = ['hours_studied', 'lms_activity_hours', 'attendance_percent', 'previous_scores']
X = df[FEATURES]
y_class = df['risk_level']
y_reg = df['exam_score']

# разбиение
X_train, X_test, y_class_train, y_class_test, y_reg_train, y_reg_test = train_test_split(
    X, y_class, y_reg, test_size=0.2, random_state=42, stratify=y_class
)

scaler = StandardScaler()
X_train_s = scaler.fit_transform(X_train)
X_test_s = scaler.transform(X_test)

results = []

# 1. Random Forest
print("1. Random Forest")
rf_clf = RandomForestClassifier(n_estimators=500, class_weight='balanced', random_state=42, n_jobs=-1)
rf_clf.fit(X_train_s, y_class_train)
pred_clf = rf_clf.predict(X_test_s)
acc_rf = accuracy_score(y_class_test, pred_clf)
recall_rf = recall_score(y_class_test, pred_clf, average=None)[0]

rf_reg = RandomForestRegressor(n_estimators=500, random_state=42, n_jobs=-1)
rf_reg.fit(X_train_s, y_reg_train)
pred_reg = rf_reg.predict(X_test_s)
r2_rf = r2_score(y_reg_test, pred_reg)
mae_rf = mean_absolute_error(y_reg_test, pred_reg)
mse_rf = mean_squared_error(y_reg_test, pred_reg)

results.append({"Модель": "Random Forest", "Acc": acc_rf, "Recall(риск)": recall_rf,
                "R²": r2_rf, "MAE": mae_rf, "MSE": mse_rf})
print(f"   Acc: {acc_rf:.4f} | Recall_риск: {recall_rf:.4f} | R²: {r2_rf:.4f} | MAE: ±{mae_rf:.2f} | MSE: {mse_rf:.4f}")

# 2. KNeighbors
print("2. KNeighbors")
knn_clf = KNeighborsClassifier(n_neighbors=5)
knn_clf.fit(X_train_s, y_class_train)
acc_knn = accuracy_score(y_class_test, knn_clf.predict(X_test_s))
recall_knn = recall_score(y_class_test, knn_clf.predict(X_test_s), average=None)[0]

knn_reg = KNeighborsRegressor(n_neighbors=5)
knn_reg.fit(X_train_s, y_reg_train)
pred_reg = knn_reg.predict(X_test_s)
r2_knn = r2_score(y_reg_test, pred_reg)
mae_knn = mean_absolute_error(y_reg_test, pred_reg)
mse_knn = mean_squared_error(y_reg_test, pred_reg)

results.append({"Модель": "KNeighbors", "Acc": acc_knn, "Recall(риск)": recall_knn,
                "R²": r2_knn, "MAE": mae_knn, "MSE": mse_knn})
print(f"   Acc: {acc_knn:.4f} | Recall_риск: {recall_knn:.4f} | R²: {r2_knn:.4f} | MAE: ±{mae_knn:.2f} | MSE: {mse_knn:.4f}")

# 3. GaussianNB
print("3. GaussianNB")
nb = GaussianNB()
nb.fit(X_train_s, y_class_train)
pred_nb = nb.predict(X_test_s)
acc_nb = accuracy_score(y_class_test, pred_nb)
recall_nb = recall_score(y_class_test, pred_nb, average=None)[0]

results.append({"Модель": "GaussianNB", "Acc": acc_nb, "Recall(риск)": recall_nb,
            "R²": "—", "MAE": "—", "MSE": "—"})
print(f"   Acc: {acc_nb:.4f} | Recall_риск: {recall_nb:.4f}")

# 4. LinearRegression
print("4. LinearRegression")
lr = LinearRegression()
lr.fit(X_train_s, y_reg_train)
pred_lr = lr.predict(X_test_s)
r2_lr = r2_score(y_reg_test, pred_lr)
mae_lr = mean_absolute_error(y_reg_test, pred_lr)
mse_lr = mean_squared_error(y_reg_test, pred_lr)

results.append({"Модель": "LinearRegression", "Acc": "—", "Recall(риск)": "—",
                "R²": r2_lr, "MAE": mae_lr, "MSE": mse_lr})
print(f"   R²: {r2_lr:.4f} | MAE: ±{mae_lr:.2f} | MSE: {mse_lr:.4f}")

# 5. АНСАМБЛЬ
print("\n5. Ансамбль (Voting)")
ensemble_clf = VotingClassifier([
    ('rf', RandomForestClassifier(n_estimators=500, class_weight='balanced', random_state=42, n_jobs=-1)),
    ('nb', GaussianNB()),
    ('knn', KNeighborsClassifier(n_neighbors=5))
], voting='soft')
ensemble_clf.fit(X_train_s, y_class_train)
pred_ens_clf = ensemble_clf.predict(X_test_s)
acc_ens = accuracy_score(y_class_test, pred_ens_clf)
recall_ens = recall_score(y_class_test, pred_ens_clf, average=None)[0]

ensemble_reg = VotingRegressor([
    ('rf', RandomForestRegressor(n_estimators=500, random_state=42, n_jobs=-1)),
    ('knn', KNeighborsRegressor(n_neighbors=5)),
    ('lr', LinearRegression())
], weights=[0.6, 0.2, 0.2])
ensemble_reg.fit(X_train_s, y_reg_train)
pred_ens_reg = ensemble_reg.predict(X_test_s)
r2_ens = r2_score(y_reg_test, pred_ens_reg)
mae_ens = mean_absolute_error(y_reg_test, pred_ens_reg)
mse_ens = mean_squared_error(y_reg_test, pred_ens_reg)

results.append({"Модель": "Ансамбль (Voting)", "Acc": acc_ens, "Recall(риск)": recall_ens,
                "R²": r2_ens, "MAE": mae_ens, "MSE": mse_ens})
print(f"   Acc: {acc_ens:.4f} | Recall_риск: {recall_ens:.4f} | R²: {r2_ens:.4f} | MAE: ±{mae_ens:.2f} | MSE: {mse_ens:.4f}")

hypotheses = [
    ("hyp1", "Все признаки", ['hours_studied','lms_activity_hours','attendance_percent','previous_scores']),
    ("hyp2", "LMS + Посещаемость", ['lms_activity_hours','attendance_percent']),
    ("hyp3", "LMS + Посещ. + Прошлые", ['lms_activity_hours','attendance_percent','previous_scores'])
]


print("ПРОВЕРКА ГИПОТЕЗ — только ансамбль")
hypothesis_results = []
best_r2 = r2_ens
best_pred = pred_ens_reg
best_real = y_reg_test
best_name = "Ансамбль (все признаки)"

for code, name, feats in hypotheses:
    print(f"\n→ {code}: {name} ({len(feats)} признаков)")
    
    X_h = df[feats]
    X_tr, X_te, y_c_tr, y_c_te, y_r_tr, y_r_te = train_test_split(
        X_h, y_class, y_reg, test_size=0.2, random_state=42, stratify=y_class
    )
    
    scaler_h = StandardScaler()
    X_tr_s = scaler_h.fit_transform(X_tr)
    X_te_s = scaler_h.transform(X_te)
    
    ens_reg = VotingRegressor([
        ('rf', RandomForestRegressor(n_estimators=500, random_state=42, n_jobs=-1)),
        ('knn', KNeighborsRegressor(n_neighbors=5)),
        ('lr', LinearRegression())
    ], weights=[0.6, 0.2, 0.2])
    
    ens_reg.fit(X_tr_s, y_r_tr)
    pred_h = ens_reg.predict(X_te_s)
    
    r2_h = r2_score(y_r_te, pred_h)
    mae_h = mean_absolute_error(y_r_te, pred_h)
    mse_h = mean_squared_error(y_r_te, pred_h)
    
    hypothesis_results.append({
        "Гипотеза": code,
        "Описание": name,
        "Признаков": len(feats),
        "R²": round(r2_h, 4),
        "MAE": round(mae_h, 2),
        "MSE": round(mse_h, 4)
    })
    
    print(f"   R²: {r2_h:.4f} | MAE: ±{mae_h:.2f} | MSE: {mse_h:.4f}")
    
    if r2_h > best_r2:
        best_r2 = r2_h
        best_pred = pred_h
        best_real = y_r_te
        best_name = f"{code}: {name}"

# Графики
plt.figure(figsize=(10, 8))
plt.scatter(best_real, best_pred, alpha=0.7, color='#9b59b6', s=80, edgecolor='white')
plt.plot([40, 100], [40, 100], 'r--', lw=2.5)
plt.title(f"ЛУЧШАЯ → {best_name}\nR² = {best_r2:.4f}", fontsize=14)
plt.xlabel("Реальный балл")
plt.ylabel("Предсказанный балл")
plt.grid(alpha=0.3)
plt.tight_layout()

graph_path = ROOT / "model" / "final_report" / "График_лучшая_гипотеза.png"
graph_path.parent.mkdir(parents=True, exist_ok=True)
plt.savefig(graph_path, dpi=300, bbox_inches='tight')
plt.close()

# Excel — два листа
excel_path = ROOT / "model" / "final_report" / "Сравнение_всех_моделей.xlsx"

wb = Workbook()
wb.remove(wb.active)

ws1 = wb.create_sheet("Все модели")
for r in dataframe_to_rows(pd.DataFrame(results).round(4), index=False, header=True):
    ws1.append(r)

ws2 = wb.create_sheet("Гипотезы")
for r in dataframe_to_rows(pd.DataFrame(hypothesis_results), index=False, header=True):
    ws2.append(r)

# Стили
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(border_style="thin")

for ws in wb.worksheets:
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

wb.save(excel_path)

print(f"   Проверено 3 гипотезы")
print(f"   Excel: {excel_path}")
print(f"   График лучшей модели: {graph_path}")
